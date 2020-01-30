import numpy as np
import pickle
import time
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import pickle
import pandas as pd
from sklearn.ensemble import RandomForestRegressor

#читаем данные из excel
work_books_8_18 = []
for i in range(15):
    print(i)
    string = 'data_result/journals-8-18-({}).xlsx'.format(i)
    wb = openpyxl.load_workbook(string)
    work_books_8_18.append(wb)

work_books_9_19 = []
for i in range(0, 15):
    print(i)
    string = 'data_result/journals-9-19-({}).xlsx'.format(i)
    wb = openpyxl.load_workbook(string)
    work_books_9_19.append(wb)


def to_float(n):# просто функция для определения является ли значение числом
    try:
        n = float(n)
    except:
        n = 'nan'
    return n


def execute(wb, sheet_name):# функция извлечения данных из excel
    # загружаем базу
    sheets = wb.get_sheet_names()
    # загружаем таблицу
    sheet = wb.get_sheet_by_name(sheet_name)
    # достаем класс и предмет
    sub = (sheet.cell(row=41, column=21).value).split(',  ')
    # создаем словари для хранения выуженной информации об учениках
    mid = dict()
    num = dict()
    exam = dict()
    fin = dict()
    absent = dict()
    num_5 = dict()
    num_4 = dict()
    num_3 = dict()
    num_2 = dict()
    for i in range(1, 35):
        if to_float(sheet.cell(row=i, column=1).value) != 'nan':
            mid.update({sheet.cell(row=i, column=2).value: 0})
            num.update({sheet.cell(row=i, column=2).value: 0})
            fin.update({sheet.cell(row=i, column=2).value: 0})
            absent.update({sheet.cell(row=i, column=2).value: 0})
            num_5.update({sheet.cell(row=i, column=2).value: 0})
            num_4.update({sheet.cell(row=i, column=2).value: 0})
            num_3.update({sheet.cell(row=i, column=2).value: 0})
            num_2.update({sheet.cell(row=i, column=2).value: 0})
            exam.update({sheet.cell(row=i, column=2).value: False})
    # print(sub[1], sub[0])
    # вытаскиваем ряды хранящие информацию о дате или типе оценки
    main = []
    for i in range(1, 501):
        if sheet.cell(row=i, column=1).value == '№':
            main.append(i + 1)
    main.append(10000000)
    n = 0
    # основная часть
    for i in range(3, 501):
        # смотрим где хранится информация о типе оценки
        if i >= main[n + 1]:
            n += 1
        main1 = main[n]
        # проверяем, стоит ли рассматривать эту строку
        if str(sheet.cell(row=i, column=2).value) != 'None':
            mark = 0
            k = 0
            fin_mark = 0
            fin_k = 0
            abse = 0
            m_5 = 0
            m_4 = 0
            m_3 = 0
            m_2 = 0
            # перебираем все оценки для ученика
            for j in range(3, 20):
                # проверяем, стоит ли рассматривать эту запись
                if sheet.cell(row=i, column=j).value == 'н':
                    abse += 1
                if to_float((sheet.cell(row=i, column=j).value)) != 'nan':
                    # проверяем, стоит ли рассматривать эту запись
                    if to_float(sheet.cell(row=main1, column=j).value) != 'nan':
                        if to_float((sheet.cell(row=i, column=j).value)) > 5:
                            pass
                        else:
                            if to_float((sheet.cell(row=i, column=j).value)) == 5:
                                m_5 += 1
                            elif to_float((sheet.cell(row=i, column=j).value)) == 4:
                                m_4 += 1
                            elif to_float((sheet.cell(row=i, column=j).value)) == 3:
                                m_3 += 1
                            elif to_float((sheet.cell(row=i, column=j).value)) == 2:
                                m_2 += 1
                            mark += to_float((sheet.cell(row=i, column=j).value))
                            k += 1
                    else:
                        fin_mark += to_float((sheet.cell(row=i, column=j).value))
                        fin_k += 1
                    # смотрим где хранится информация об экзамене
                    if sheet.cell(row=main1, column=j).value == 'Э':
                        exam.update({sheet.cell(row=i, column=2).value: int(sheet.cell(row=i, column=j).value)})
            # заполняем словари
            if k != 0:
                middle = mark / k
                if to_float((sheet.cell(row=i, column=1).value)) != 'nan':
                    # добавляем и обновляем среднюю оценку
                    if mid[sheet.cell(row=i, column=2).value] != 0:
                        mid.update(
                            {sheet.cell(row=i, column=2).value: (mid[sheet.cell(row=i, column=2).value] + middle) / 2})
                    else:
                        mid.update({sheet.cell(row=i, column=2).value: (middle)})
                    # добавляем и обновляем количество оценок
                    num.update({sheet.cell(row=i, column=2).value: num[sheet.cell(row=i, column=2).value] + k})
            if fin_k != 0:
                middle = fin_mark / fin_k
                if to_float((sheet.cell(row=i, column=1).value)) != 'nan':
                    if fin[sheet.cell(row=i, column=2).value] != 0:
                        fin.update(
                            {sheet.cell(row=i, column=2).value: (fin[sheet.cell(row=i, column=2).value] + middle) / 2})
                    else:
                        fin.update({sheet.cell(row=i, column=2).value: (middle)})
            if sheet.cell(row=i, column=2).value != 'Обучающийся':
                absent.update({sheet.cell(row=i, column=2).value: absent[sheet.cell(row=i, column=2).value] + abse})
                num_5.update({sheet.cell(row=i, column=2).value: num_5[sheet.cell(row=i, column=2).value] + m_5})
                num_4.update({sheet.cell(row=i, column=2).value: num_4[sheet.cell(row=i, column=2).value] + m_4})
                num_3.update({sheet.cell(row=i, column=2).value: num_3[sheet.cell(row=i, column=2).value] + m_3})
                num_2.update({sheet.cell(row=i, column=2).value: num_2[sheet.cell(row=i, column=2).value] + m_2})
    return mid, num, exam, sub[1], sub[0], fin, absent, num_5, num_4, num_3, num_2


def to_dicts(wb, sheet_name):# перевод данных в тип словарь
    mid, num, exam, sub, class_nom, fin, absent, num_5, num_4, num_3, num_2 = execute(wb, sheet_name)
    d1 = dict()
    for i in mid:
        d1.update({i: [mid[i], num[i], fin[i], absent[i], num_5[i], num_4[i], num_3[i], num_2[i]]})
    return d1, sub, class_nom, mid


def get_sub(sheet_name, wb):# получить предмет из таблицы
    try:
        sheet = wb[sheet_name]
        sub = sheet.cell(row=41, column=21).value.split(',  ')
        return sub[0], sub[1]
    except:
        return 0, 0

def is_examing(a, subs):# проверка сдается ли по данному предмету экзамен
    for i in subs:
        if a.lower() in i.lower():
            return True, subs[i]
        elif 'обществознание' in a.lower():
            return True, 'social'
        elif 'история' in a.lower() and 'россии' not in a.lower():
            return True, 'hist'
    return False, 0


def results(work_book):# извлечение результатов экзаменов из excel
    subs = {'Алгебра': 'alg', 'Геометрия': 'geom', 'Русский язык': 'ru', 'Физика': 'phis', 'Химия': 'chem',
            'Информатика и ИКТ': 'inf', 'Биология': 'bio',
            'География': 'geo', 'История': 'hist', 'Обществознание': 'social',
            'Литература': 'lit', 'Английский язык': 'eng'}
    res = dict()
    for i in subs:
        res.update({subs[i]: dict()})
    sheets = work_book.get_sheet_names()
    for j in sheets:
        a = get_sub(j, work_book)[1]
        if a != 0:
            if is_examing(str(a), subs)[0]:
                d, sub, _, _ = to_dicts(work_book, j)
                res[is_examing(str(a), subs)[1]].update(d)
    return res, subs


def get_keys(d):# функция извлечения ключей из словарей
    a = []
    for i in d:
        a.append(i)
    return a


def get_students(work_book):# получаем список студентов
    a = []
    res, subs = results(work_book)
    for i in subs:
        df = get_keys(res[subs[i]])
        a.append(set(df))
    s = a[0]
    for i in a:
        s = s & i
    return s


def prepareing(work_book):# формирование словаря значений
    res, subs = results(work_book)
    s = get_students(work_book)
    di = dict()
    for i in s:
        di.update({i: list()})
    for i in subs:
        model1 = pickle.load(open(r'C:\Users\artem\Data2020\models/{}_5s.sav'.format(subs[i]), 'rb'))
        model2 = pickle.load(open(r'C:\Users\artem\Data2020\models/{}_4s.sav'.format(subs[i]), 'rb'))
        model3 = pickle.load(open(r'C:\Users\artem\Data2020\models/{}_3s.sav'.format(subs[i]), 'rb'))
        model5 = pickle.load(open(r'C:\Users\artem\Data2020\models/{}_absent.sav'.format(subs[i]), 'rb'))
        model6 = pickle.load(open(r'C:\Users\artem\Data2020\models/fin_{}.sav'.format(subs[i]), 'rb'))
        model7 = pickle.load(open(r'C:\Users\artem\Data2020\models/num_{}.sav'.format(subs[i]), 'rb'))
        g_1 = []
        g_2 = []
        g_3 = []
        g_4 = []
        g_5 = []
        g_6 = []
        js = []
        for j in s:
            g = res[subs[i]][j]
            g_1.append(g[1])
            g_2.append(g[2])
            g_3.append(g[3])
            g_4.append(g[4])
            g_5.append(g[5])
            g_6.append(g[6])
            js.append(j)
        g_1_pred = model7.predict(np.array(g_1).reshape(-1, 1))
        g_2_pred = model6.predict(np.array(g_2).reshape(-1, 1))
        g_3_pred = model5.predict(np.array(g_3).reshape(-1, 1))
        g_4_pred = model1.predict(np.array(g_4).reshape(-1, 1))
        g_5_pred = model2.predict(np.array(g_5).reshape(-1, 1))
        g_6_pred = model3.predict(np.array(g_6).reshape(-1, 1))
        for i in range(len(g_1)):
            if g_1[i] != 0:
                marks = int(g_1[i]+g_1_pred[i])
                s3 = int(g_6[i]+g_6_pred[i])
                s4 = int(g_5[i]+g_5_pred[i])
                s5 = int(g_4[i]+g_4_pred[i])
                h = (3 * s3 + 4 * s4 + 5 * s5) / (5 * marks) * 100
                abse = int(g_3[i]+g_3_pred[i])
                avg = int(g_2[i]+g_2_pred[i])
                a = di[js[i]]
                a.extend([marks, h, abse, avg])
            else:
                marks = int(g_1[i]+g_1_pred[i])
                abse = int(g_3[i]+g_3_pred[i])
                avg = int(g_2[i]+g_2_pred[i])
                a = di[js[i]]
                a.extend([marks, 0, abse, avg])
            di.update({js[i]: a})
    return di


def prediction(work_book):# функция обработки предсказанного результата
    res = prepareing(work_book)
    a = []
    for i in res:
        ret = predicted(res[i])
        a.append([i, ret])
    return a


def predicted(X):# функция предсказания
    a = []
    ret = []
    exam = ['exam_phis', 'exam_chem', 'exam_inf', 'exam_bio', 'exam_geo', 'exam_hist', 'exam_social', 'exam_lit',
            'exam_eng']
    exam_2 = ['exam_alg', 'exam_ru']
    for sub in exam_2:
        model1 = pickle.load(open(r'C:\Users\artem\models/prob_{}_mark.sav'.format(sub), 'rb'))
        b1 = model1.predict([X])
        ret.append([1.0,sub[5:], b1[0]])
    for i in range(len(X)):
        a.append(list())
    for sub in exam:
        model = pickle.load(open(r'C:\Users\artem\models/prob_{}.sav'.format(sub), 'rb'))
        model1 = pickle.load(open(r'C:\Users\artem\models/prob_{}_mark.sav'.format(sub), 'rb'))
        b2 = model.predict([X])
        b = model.predict_proba([X])
        b1 = model1.predict([X])
        for i in range(len(b)):
            if b2[i] == 'True':
                h = max(b[i][1], b[i][0])
            else:
                h = min(b[i][1], b[i][0])
            a[i].append((h, sub[5:], b1[i]))
    g = []
    for i in range(len(a)):
        g.append(sorted(a[i], key=lambda x: x[0]))
    for i in g:
        exams = i[-2:]
        for j in exams:
            ret.append([j[0], j[1], j[2]])
    return ret

# далее собираем данные в csv:

subs = {'Алгебра': 'alg', 'Геометрия': 'geom', 'Русский язык': 'ru', 'Физика': 'phis', 'Химия': 'chem',
        'Информатика и ИКТ': 'inf', 'Биология': 'bio',
        'География': 'geo', 'История (включая москвоведение)': 'hist', 'Обществознание (включая экономику и право)': 'social',
        'Литература': 'lit', 'Английский язык': 'eng'}
import csv
def to_csv(filename):
    with open(filename, "w", newline="") as file:
        columns = ['name', 'avg', 'marks', 'exam', 'fin_avg', 'absent', 'num_5', 'num_4', 'num_3', 'num_2']
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
    return 0

def to_list(wb, sheet_name):
    mid, num, exam, sub, class_nom,  fin, absent, num_5, num_4, num_3, num_2 = execute(wb, sheet_name)
    d = []
    for i in mid:
        d.append([i, mid[i], num[i], exam[i], fin[i], absent[i], num_5[i], num_4[i], num_3[i], num_2[i]])
    return d, sub,  class_nom


def add_to_csv(wb, filename, sheet_name):
    students, _, _ = to_list(wb, sheet_name)
    with open(filename, "a", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(students)
    return 0
for i in subs:
    to_csv('data_csv/' + str(subs[i]) + '_9.csv')

for wb in work_books_9_19:
    sheets = wb.get_sheet_names()
    for j in sheets:
        a = get_sub(j, wb)[1]
        try:
            if is_examing(str(a), subs)[0]:
                file = is_examing(str(a), subs)[1]
                filename = 'data_csv/' + str(file) + '_9.csv'
                add_to_csv(wb, filename, j)
        except:
            pass

# обучаем модели на прогноз разницы

for i in subs:
    if subs[i]!='hist':
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        d_1 = dict()
        d_2 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['marks'][k]})
        for k in range(286):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: df1['marks'][k]})
    else:
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        df2 = pd.read_csv('data_csv/ru_hist_9.csv')
        d_1 = dict()
        d_2 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['marks'][k]})
        for k in range(286):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: (df1['marks'][k]+df2['marks'][k])/2})
    X = []
    y = []
    y1 = []
    for j in d_1:
        try:
            y.append(d_2[j])
            y1.append(float(d_1[j]))
            X.append(j)
        except:
            pass
    dif=0
    diff = []
    diff_1 = []
    for l in range(len(y)):
        dif+=abs(y[l]-y1[l])
        diff.append((y1[l], (y[l]-y1[l])))
        diff_1.append(abs(y[l]-y1[l]))
    diff_sorted = sorted(diff, key = lambda x: x[0])
    print(subs[i], dif/sum(y), dif/len(X))
    plt.plot(X, y)
    plt.plot(X, y1)
    plt.xlabel('Students \n {}, {}, {},\n {}, {}'.format(i, dif/(5*len(y)), dif/len(X), max(diff_1), min(diff_1)))
    plt.ylabel('Avg. Mark')
    plt.show()
    X_train = []
    y_train = []
    for f in diff_sorted:
        X_train.append(f[0])
        y_train.append(f[1])
    plt.plot(X_train, y_train)
    model5 = RandomForestRegressor(max_depth=1000, random_state=0, n_estimators = 1000).fit(np.array(X_train).reshape(-1, 1), y_train) # SVR(C = 100000, gamma = 'auto').fit(np.array(X).reshape(-1, 1), y)
    y_1 = model5.predict(np.array(X_train).reshape(-1, 1))
    plt.plot(X_train, y_1)
    plt.xlabel('Оценка')
    plt.ylabel('Корректировка')
    plt.show()
    pickle.dump(model5, open('models/num_{}.sav'.format(subs[i]), 'wb'))

for i in subs:
    if subs[i]!='hist':
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        d_1 = dict()
        d_2 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['fin_avg'][k]})
        for k in range(286):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: df1['fin_avg'][k]})
    else:
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        df2 = pd.read_csv('data_csv/ru_hist_9.csv')
        d_1 = dict()
        d_2 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['fin_avg'][k]})
        for k in range(286):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: (df1['fin_avg'][k]+df2['fin_avg'][k])/2})
    X = []
    y = []
    y1 = []
    for j in d_1:
        try:
            y.append(d_2[j])
            y1.append(float(d_1[j]))
            X.append(j)
        except:
            pass
    dif=0
    diff = []
    diff_1 = []
    for l in range(len(y)):
        dif+=abs(y[l]-y1[l])
        diff.append((y1[l], (y[l]-y1[l])))
        diff_1.append(abs(y[l]-y1[l]))
    diff_sorted = sorted(diff, key = lambda x: x[0])
    print(subs[i], dif/sum(y), dif/len(X))
    plt.plot(X, y)
    plt.plot(X, y1)
    plt.xlabel('Students \n {}, {}, {},\n {}, {}'.format(i, dif/(5*len(y)), dif/len(X), max(diff_1), min(diff_1)))
    plt.ylabel('Avg. Mark')
    plt.show()
    X_train = []
    y_train = []
    for f in diff_sorted:
        X_train.append(f[0])
        y_train.append(f[1])
    plt.plot(X_train, y_train)
    model5 = RandomForestRegressor(max_depth=1000, random_state=0, n_estimators = 1000).fit(np.array(X_train).reshape(-1, 1), y_train) # SVR(C = 100000, gamma = 'auto').fit(np.array(X).reshape(-1, 1), y)
    y_1 = model5.predict(np.array(X_train).reshape(-1, 1))
    plt.plot(X_train, y_1)
    plt.xlabel('Оценка')
    plt.ylabel('Корректировка')
    plt.show()
    pickle.dump(model5, open('models/fin_{}.sav'.format(subs[i]), 'wb'))

for i in subs:
    if subs[i]!='hist':
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        d_1 = dict()
        d_2 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['num_5'][k]})
        for k in range(len(df1['name'])):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: df1['num_5'][k]})
    else:
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        df2 = pd.read_csv('data_csv/ru_hist_9.csv')
        d_1 = dict()
        d_2 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['num_5'][k]})
        for k in range(len(df1['name'])):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: (df1['num_5'][k]+df2['num_5'][k])/2})
    X = []
    y = []
    y1 = []
    for j in d_1:
        try:
            y.append(d_2[j])
            y1.append(float(d_1[j]))
            X.append(j)
        except:
            pass
    dif=0
    diff = []
    diff_1 = []
    for l in range(len(y)):
        dif+=abs(y[l]-y1[l])
        diff.append((y1[l], (y[l]-y1[l])))
        diff_1.append(abs(y[l]-y1[l]))
    diff_sorted = sorted(diff, key = lambda x: x[0])
    print(subs[i], dif/sum(y), dif/len(X))
    plt.plot(X, y)
    plt.plot(X, y1)
    plt.xlabel('Students \n {}, {}, {},\n {}, {}'.format(i, dif/(5*len(y)), dif/len(X), max(diff_1), min(diff_1)))
    plt.ylabel('Avg. Mark')
    plt.show()
    X_train = []
    y_train = []
    for f in diff_sorted:
        X_train.append(f[0])
        y_train.append(f[1])
    plt.plot(X_train, y_train)
    model5 = RandomForestRegressor(max_depth=1000, random_state=0, n_estimators = 1000).fit(np.array(X_train).reshape(-1, 1), y_train) # SVR(C = 100000, gamma = 'auto').fit(np.array(X).reshape(-1, 1), y)
    y_1 = model5.predict(np.array(X_train).reshape(-1, 1))
    plt.plot(X_train, y_1)
    plt.xlabel('Оценка')
    plt.ylabel('Корректировка')
    plt.show()
    pickle.dump(model5, open('models/{}_5s.sav'.format(subs[i]), 'wb'))
for h in range(2, 6):
    for i in subs:
        if subs[i]!='hist':
            df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
            df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
            d_1 = dict()
            d_2 = dict()
            for k in range(len(df['name'])):
                if df['avg'][k]!=0:
                    d_1.update({df['name'][k]: df['num_{}'.format(h)][k]})
            for k in range(len(df1['name'])):
                if df1['avg'][k]!=0:
                    d_2.update({df1['name'][k]: df1['num_{}'.format(h)][k]})
        else:
            df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
            df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
            df2 = pd.read_csv('data_csv/ru_hist_9.csv')
            d_1 = dict()
            d_2 = dict()
            for k in range(len(df['name'])):
                if df['avg'][k]!=0:
                    d_1.update({df['name'][k]: df['num_{}'.format(h)][k]})
            for k in range(len(df1['name'])):
                if df1['avg'][k]!=0:
                    d_2.update({df1['name'][k]: (df1['num_{}'.format(h)][k]+df2['num_{}'.format(h)][k])/2})
        X = []
        y = []
        y1 = []
        for j in d_1:
            try:
                y.append(d_2[j])
                y1.append(float(d_1[j]))
                X.append(j)
            except:
                pass
        dif=0
        diff = []
        diff_1 = []
        for l in range(len(y)):
            dif+=abs(y[l]-y1[l])
            diff.append((y1[l], (y[l]-y1[l])))
            diff_1.append(abs(y[l]-y1[l]))
        diff_sorted = sorted(diff, key = lambda x: x[0])
        print(subs[i], dif/sum(y), dif/len(X))
        plt.plot(X, y)
        plt.plot(X, y1)
        plt.xlabel('Students \n {}, {}, {},\n {}, {}'.format(i, dif/(5*len(y)), dif/len(X), max(diff_1), min(diff_1)))
        plt.ylabel('Avg. Mark')
        plt.show()
        X_train = []
        y_train = []
        for f in diff_sorted:
            X_train.append(f[0])
            y_train.append(f[1])
        plt.plot(X_train, y_train)
        model5 = RandomForestRegressor(max_depth=1000, random_state=0, n_estimators = 1000).fit(np.array(X_train).reshape(-1, 1), y_train) # SVR(C = 100000, gamma = 'auto').fit(np.array(X).reshape(-1, 1), y)
        y_1 = model5.predict(np.array(X_train).reshape(-1, 1))
        plt.plot(X_train, y_1)
        plt.xlabel('Оценка')
        plt.ylabel('Корректировка')
        plt.show()
        pickle.dump(model5, open('models/{}_{}s.sav'.format(subs[i], h), 'wb'))

for i in subs:
    if subs[i]!='hist':
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        d_1 = dict()
        d_2 = dict()
        d_1_1 = dict()
        d_2_1 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['marks'][k]})
                d_1_1.update({df['name'][k]: df['avg'][k]})
        for k in range(len(df1['name'])):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: df1['marks'][k]})
                d_2_1.update({df1['name'][k]: df1['avg'][k]})
    else:
        df = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
        df1 = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
        df2 = pd.read_csv('data_csv/ru_hist_9.csv')
        d_1 = dict()
        d_2 = dict()
        d_1_1 = dict()
        d_2_1 = dict()
        for k in range(len(df['name'])):
            if df['avg'][k]!=0:
                d_1.update({df['name'][k]: df['marks'][k]})
                d_1_1.update({df['name'][k]: df['avg'][k]})
        for k in range(len(df1['name'])):
            if df1['avg'][k]!=0:
                d_2.update({df1['name'][k]: df1['marks'][k]})
                d_2_1.update({df1['name'][k]: (df1['avg'][k]+df2['avg'][k])/2})
    X = []
    y = []
    y1 = []
    y_1 = []
    y1_1 = []
    for j in d_1:
        try:
            y.append(d_2[j])
            y1.append(float(d_1[j]))
            y_1.append(d_2_1[j])
            y1_1.append(float(d_1_1[j]))
            X.append(j)
        except:
            pass
    dif=0
    diff = []
    diff_1 = []
    for l in range(len(y)):
        dif+=abs(y[l]-y1[l])
        diff.append((y1[l], (y[l]-y1[l])))
        diff_1.append(abs(y[l]-y1[l]))
    diff_sorted = sorted(diff, key = lambda x: x[0])
    print(subs[i], dif/sum(y), dif/len(X))
    plt.plot(X, y)
    plt.plot(X, y1)
    plt.xlabel('Students \n {}, {}, {},\n {}, {}'.format(i, dif/(5*len(y)), dif/len(X), max(diff_1), min(diff_1)))
    plt.ylabel('Avg. Mark')
    plt.show()
    diff = []
    for l in range(len(y)):
        diff.append((y1_1[l], (y[l]-y1[l])))
    diff_sorted = sorted(diff, key = lambda x: x[0])
    X_train = []
    y_train = []
    for f in diff_sorted:
        X_train.append(f[0])
        y_train.append(f[1])
    plt.plot(X_train, y_train)
    model5 = RandomForestRegressor(max_depth=1000, random_state=0, n_estimators = 1000).fit(np.array(X_train).reshape(-1, 1), y_train) # SVR(C = 100000, gamma = 'auto').fit(np.array(X).reshape(-1, 1), y)
    y_1 = model5.predict(np.array(X_train).reshape(-1, 1))
    plt.plot(X_train, y_1)
    plt.xlabel('Оценка')
    plt.ylabel('Корректировка')
    plt.show()
    pickle.dump(model5, open('models_1/{}.sav'.format(subs[i]), 'wb'))

# приступаем к обучению финальных моделей

res = dict()
for i in subs:
    res.update({subs[i]: dict()})
for wb in work_books_9_19:
    sheets = wb.get_sheet_names()
    for j in sheets:
        a = get_sub(j, wb)[1]
        if a!=0:
            if is_examing(a, subs)[0]:
                d, sub, _, _ = to_dicts(wb, j)
                print(is_examing(a, subs)[1], sub)
                res[is_examing(a, subs)[1]].update(d)
a = []
import pandas as pd
for i in subs:
    df = pd.read_csv('data_csv/{}_9.csv'.format(subs[i]))
    df1 = pd.read_csv('data_csv/{}_8.csv'.format(subs[i]))
    a.append(set(df.name) & set(df1.name))
    print(i)
s = a[0]
for i in a:
    s = s & i
di = dict()
for i in s:
    di.update({i: list()})
    #print(i)
for i in subs:
    for j in s:
        g = res[subs[i]][j]
        if g[1]!=0:
            h = (3*g[6]+4*g[5]+5*g[4])/(5*g[1])*100
            abse = g[3]
            avg = g[2]
            marks = g[1]
            #print(di[j], j)
            a = di[j]
            a.extend([ marks, h, abse, avg])
        else:
            abse = g[3]
            avg = g[2]
            marks = g[1]
            a = di[j]
            a.extend([marks, 0, abse, avg])
        di.update({j: a})
    print(i)
for i in subs:
    for j in s:
        g = res[subs[i]][j]
        a = di[j]
        a.extend([g[-1]])
        di.update({j: a})
d = []
for i in di:
    d.append(di[i])
columns = ['num_alg', 'abs_alg','ma_alg','avg_alg', 'num_geom', 'abs_geom','ma_geom','avg_geom', 'num_ru', 'abs_ru','ma_ru','avg_ru', 'num_phis', 'abs_phis','ma_phis','avg_phis', 'num_chem', 'abs_chem','ma_chem','avg_chem', 'num_inf', 'abs_inf','ma_inf','avg_inf', 'num_bio', 'abs_bio','ma_bio','avg_bio', 'num_geo', 'abs_geo','ma_geo','avg_geo', 'num_hist', 'abs_hist','ma_hist','avg_hist', 'num_social', 'abs_social','ma_social','avg_social', 'num_lit', 'abs_lit','ma_lit','avg_lit', 'num_eng', 'abs_eng','ma_eng','avg_eng']
exam = ['exam_phis', 'exam_chem', 'exam_inf', 'exam_bio', 'exam_geo', 'exam_hist', 'exam_social', 'exam_lit', 'exam_eng']
import csv
with open('for_val_and_train.scv', "w", newline="") as file:
        columns = ['num_alg', 'abs_alg','ma_alg','avg_alg', 'num_geom', 'abs_geom','ma_geom','avg_geom', 'num_ru', 'abs_ru','ma_ru','avg_ru', 'num_phis', 'abs_phis','ma_phis','avg_phis', 'num_chem', 'abs_chem','ma_chem','avg_chem', 'num_inf', 'abs_inf','ma_inf','avg_inf', 'num_bio', 'abs_bio','ma_bio','avg_bio', 'num_geo', 'abs_geo','ma_geo','avg_geo', 'num_hist', 'abs_hist','ma_hist','avg_hist', 'num_social', 'abs_social','ma_social','avg_social', 'num_lit', 'abs_lit','ma_lit','avg_lit', 'num_eng', 'abs_eng','ma_eng','avg_eng',  'exam_alg', 'exam_geom', 'exam_ru', 'exam_phis', 'exam_chem', 'exam_inf', 'exam_bio', 'exam_geo', 'exam_hist', 'exam_social', 'exam_lit', 'exam_eng']
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
with open('for_val_and_train.scv', "a", newline="") as file:
    writer = csv.writer(file)
    writer.writerows(d)
import pickle
for sub in exam:
    df_1 = df[0:250]
    df_2 = df[200:]
    X = []
    y = []
    for i in range(250):
        a = []
        for j in exam:
            j=sub
            e = df_1[j][i]
            if e!='False':
                e = int(e)
                y.append(e)
            else:
                e = False
            a = []
            if e:
                for j in columns:
                    a.append(df_1[j][i])
                X.append(a)
    X_test = []
    y_test = []
    for i in range(200, 250):
        a = []
        for j in exam:
            j=sub
            e = df_2[j][i]
            if e!='False':
                e = int(e)
                y_test.append(e)
            else:
                e = False
            if e:
                a = []
                for j in columns:
                    a.append(df_2[j][i])
                X_test.append(a)
    from sklearn.neural_network import MLPClassifier
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.svm import SVC
    models = [
        MLPClassifier(alpha=0.1, max_iter=10000)
    ]
    for i in models:
        i.fit(X, y)
        pickle.dump(i, open('models/prob_{}_mark.sav'.format(sub), 'wb'))
        test = i.predict(X_test)
        al = 0
        yes = 0
        for j in range(len(y_test)):
            yes += abs(y_test[j] - test[j])
            al += y_test[j]
import pickle
exam = ['exam_phis', 'exam_chem', 'exam_inf', 'exam_bio', 'exam_geo', 'exam_hist', 'exam_social', 'exam_lit', 'exam_eng']
for sub in exam:
    df_1 = df[0:250]
    df_2 = df[200:]
    X = []
    for i in range(250):
        a = []
        for j in columns:
            a.append(df_1[j][i])
        X.append(a)
    y = []
    for i in range(250):
        a = []
        for j in exam:
            j=sub
            e = df_1[j][i]
            if e!='False':
                e = True
            else:
                e = False
        y.append(e)
    X_test = []
    for i in range(200, 250):
        a = []
        for j in columns:
            a.append(df_2[j][i])
        X_test.append(a)
    y_test = []
    for i in range(200, 250):
        a = []
        for j in exam:
            j=sub
            e = df_2[j][i]
            if e!='False':
                e = True
            else:
                e = False
        y_test.append(e)
    from sklearn.neural_network import MLPClassifier
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.svm import SVC
    models = [
        MLPClassifier(alpha=0.1, max_iter=10000)
    ]
    for i in models:
        i.fit(X, y)
        pickle.dump(i, open('models/prob_{}.sav'.format(sub), 'wb'))
        test = i.predict(X_test)
        al = 0
        yes = 0
        for j in range(len(y_test)):
            if y_test[j] == test[j]:
                yes += 1
            al += 1

print('Готово')